"""Tạo file Excel sổ đo thủy chuẩn hạng 4."""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from doc_mau import MauTram
from phan_bo_tram import DocTram, KetQuaTram, tinh_ket_qua_tram

INPUT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")


@dataclass
class DuAn:
    do_tu: str = ""
    den: str = ""
    bat_dau: str = ""
    ket_thuc: str = ""
    ngay_do: str = ""
    thoi_tiet: str = ""
    nguoi_do: str = ""
    nguoi_ghi: str = ""
    nguoi_kiem_tra: str = ""


@dataclass
class TongHop:
    l_sau: float
    l_truoc: float
    h_tong: float
    i_tong: float


def _station_rows(n_tram: int) -> tuple[list[int], list[int]]:
    s_rows = [13 + i * 4 for i in range(n_tram)]
    t_rows = [14 + i * 4 for i in range(n_tram)]
    return s_rows, t_rows


def _sum_formula(col: str, rows: list[int]) -> str:
    refs = ",".join(f"{col}{r}" for r in rows)
    return f"=SUM({refs})"


def _write_header(ws, du_an: DuAn) -> None:
    ws["A1"] = "SỔ ĐO THỦY CHUẨN HẠNG 4"

    ws["A3"] = "Đo từ:"
    ws["C3"] = du_an.do_tu
    ws["H3"] = "Đến:"
    ws["J3"] = du_an.den

    ws["A4"] = "Bắt đầu lúc:"
    ws["C4"] = du_an.bat_dau
    ws["H4"] = "Ngày đo:"
    ws["J4"] = du_an.ngay_do

    ws["A5"] = "Kết thúc lúc:"
    ws["C5"] = du_an.ket_thuc
    ws["H5"] = "Người đo:"
    ws["J5"] = du_an.nguoi_do

    ws["A6"] = "Thời tiết:"
    ws["C6"] = du_an.thoi_tiet
    ws["H6"] = "Người ghi:"
    ws["J6"] = du_an.nguoi_ghi

    ws["H7"] = "Người kiểm tra:"
    ws["J7"] = du_an.nguoi_kiem_tra

    labels = {
        (9, "A"): "Số",
        (9, "B"): "Mia",
        (9, "C"): "Chỉ trên",
        (9, "D"): "Mia",
        (9, "E"): "Chỉ trên",
        (9, "F"): "Ký",
        (9, "H"): "Số đọc trên mia",
        (9, "J"): "K",
        (9, "K"): "Chênh",
        (10, "A"): "trạm",
        (10, "B"): "sau",
        (10, "C"): "Chỉ dưới",
        (10, "D"): "trước",
        (10, "E"): "Chỉ dưới",
        (10, "F"): "hiệu",
        (10, "H"): "Mặt",
        (10, "I"): "Mặt",
        (10, "J"): "+ đen",
        (10, "K"): "cao",
        (10, "L"): "Ghi",
        (11, "A"): "đo",
        (11, "B"): "K/C sau",
        (11, "D"): "K/C trước",
        (11, "F"): "mia",
        (11, "H"): "đen",
        (11, "I"): "đỏ",
        (11, "J"): "- đỏ",
        (11, "K"): "trung",
        (11, "L"): "chú",
        (12, "B"): "ΔS",
        (12, "D"): "ΣΔS",
        (12, "K"): "bình",
        (12, "N"): "Hằng số mia K",
    }
    for (row, col), value in labels.items():
        ws[f"{col}{row}"] = value


def _write_station_block(
    ws,
    tram: int,
    doc: DocTram,
    kq: KetQuaTram,
    mt: MauTram,
    k1: float,
    k2: float,
    delta_k: float,
    prev_e_delta_row: int | None,
) -> int:
    r_s = 13 + (tram - 1) * 4
    r_t = r_s + 1
    r_st = r_s + 2
    r_delta = r_s + 3

    label_s = "K1" if mt.n_s == k1 else "K2"
    label_t = "K1" if mt.n_t == k1 else "K2"

    ws[f"A{r_s}"] = tram
    ws[f"F{r_s}"] = "S"
    ws[f"C{r_s}"] = doc.c_s
    ws[f"E{r_s}"] = doc.e_s
    ws[f"H{r_s}"] = doc.h_s
    ws[f"I{r_s}"] = doc.i_s
    ws[f"L{r_s}"] = label_s
    ws[f"N{r_s}"] = mt.n_s

    ws[f"F{r_t}"] = "T"
    ws[f"C{r_t}"] = doc.c_t
    ws[f"E{r_t}"] = doc.e_t
    ws[f"H{r_t}"] = doc.h_t
    ws[f"I{r_t}"] = doc.i_t
    ws[f"L{r_t}"] = label_t
    ws[f"N{r_t}"] = mt.n_t

    ws[f"J{r_s}"] = kq.j_s
    ws[f"J{r_t}"] = kq.j_t

    ws[f"C{r_st}"] = kq.kc_sau
    ws[f"E{r_st}"] = kq.kc_truoc
    ws[f"F{r_st}"] = "S-T"
    ws[f"H{r_st}"] = kq.chenh_h
    ws[f"I{r_st}"] = kq.chenh_i
    ws[f"J{r_st}"] = kq.j_st
    ws[f"K{r_st}"] = kq.chenh_cao
    ws[f"N{r_st}"] = delta_k

    ws[f"C{r_delta}"] = kq.chenh_kc
    if tram == 1:
        ws[f"E{r_delta}"] = kq.chenh_kc
    elif prev_e_delta_row is not None:
        ws[f"E{r_delta}"] = f"=E{prev_e_delta_row}+C{r_delta}"

    for r in (r_s, r_t):
        for col in ("C", "E", "H", "I"):
            ws[f"{col}{r}"].fill = INPUT_FILL

    return r_delta


def _write_summary(
    ws,
    n_tram: int,
    tong_hop: TongHop,
    s_rows: list[int],
    t_rows: list[int],
) -> None:
    r_tong_s = 13 + 4 * n_tram
    r_tong_t = 14 + 4 * n_tram
    r_user_st = 15 + 4 * n_tram
    r_chenh = 16 + 4 * n_tram
    r_f_sum_s = 18 + 4 * n_tram
    r_f_sum_t = 19 + 4 * n_tram
    r_f_calc = 20 + 4 * n_tram
    r_f_chenh = 21 + 4 * n_tram

    l_km = (tong_hop.l_sau + tong_hop.l_truoc) / 2 / 1000

    ws[f"A{r_tong_s}"] = "Tổng"
    ws[f"F{r_tong_s}"] = "S"
    for col in ("C", "E", "H", "I"):
        ws[f"{col}{r_tong_s}"] = _sum_formula(col, s_rows)

    ws[f"F{r_tong_t}"] = "T"
    for col in ("C", "E", "H", "I"):
        ws[f"{col}{r_tong_t}"] = _sum_formula(col, t_rows)

    ws[f"C{r_user_st}"] = tong_hop.l_sau
    ws[f"E{r_user_st}"] = tong_hop.l_truoc
    ws[f"F{r_user_st}"] = "S-T"
    ws[f"H{r_user_st}"] = tong_hop.h_tong
    ws[f"I{r_user_st}"] = tong_hop.i_tong
    ws[f"K{r_user_st}"] = tong_hop.h_tong

    delta_l = tong_hop.l_sau - tong_hop.l_truoc
    ws[f"C{r_chenh}"] = delta_l
    ws[f"E{r_chenh}"] = delta_l

    for col in ("C", "E", "H", "I"):
        ws[f"{col}{r_f_sum_s}"] = _sum_formula(col, s_rows)
        ws[f"{col}{r_f_sum_t}"] = _sum_formula(col, t_rows)

    ws[f"B{r_f_calc}"] = "Lđo / 2 ="
    ws[f"C{r_f_calc}"] = f"=(C{r_f_sum_s}-C{r_f_sum_t})/10"
    ws[f"E{r_f_calc}"] = f"=(E{r_f_sum_s}-E{r_f_sum_t})/10"
    ws[f"H{r_f_calc}"] = f"=H{r_f_sum_s}-H{r_f_sum_t}"
    ws[f"I{r_f_calc}"] = f"=I{r_f_sum_s}-I{r_f_sum_t}"
    ws[f"L{r_f_calc}"] = f"±20 x sqrt({l_km:.3f})"

    ws[f"B{r_f_chenh}"] = "± 0,5 x"
    ws[f"C{r_f_chenh}"] = f"=C{r_f_calc}-E{r_f_calc}"
    ws[f"E{r_f_chenh}"] = f"=C{r_f_chenh}"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["L"].width = 18


def build_workbook(
    du_an: DuAn,
    n_tram: int,
    k1: float,
    k2: float,
    delta_k: float,
    tong_hop: TongHop,
    tram_list: list[DocTram],
    mau_sel: list[MauTram],
    chenh_cao_list: list[float],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _write_header(ws, du_an)

    prev_e_delta_row: int | None = None
    for i, doc in enumerate(tram_list, start=1):
        mt = mau_sel[i - 1]
        kq = tinh_ket_qua_tram(
            doc, mt.n_s, mt.n_t, delta_k, i, chenh_cao_list[i - 1]
        )
        r_delta = _write_station_block(
            ws, i, doc, kq, mt, k1, k2, delta_k, prev_e_delta_row
        )
        prev_e_delta_row = r_delta

    s_rows, t_rows = _station_rows(n_tram)
    _write_summary(ws, n_tram, tong_hop, s_rows, t_rows)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
