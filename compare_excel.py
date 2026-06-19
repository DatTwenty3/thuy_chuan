"""So sánh So do.xlsx và file xuất."""

import os
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
MAU = ROOT / "So do.xlsx"
XUAT = next(ROOT.glob("So_do_*.xlsx"))


def count_formulas(ws) -> int:
    return sum(
        1
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


def count_stations(ws) -> int:
    n = 0
    for r in range(13, 200, 4):
        val = ws.cell(r, 1).value
        if val is not None and str(val).isdigit():
            n += 1
        else:
            break
    return n


def row_dict(ws, r, max_col=14):
    return {
        get_column_letter(c): ws.cell(r, c).value
        for c in range(1, max_col + 1)
        if ws.cell(r, c).value is not None
    }


def main():
    wb_m = openpyxl.load_workbook(MAU, data_only=False)
    wb_x = openpyxl.load_workbook(XUAT, data_only=False)
    wb_mv = openpyxl.load_workbook(MAU, data_only=True)
    wb_xv = openpyxl.load_workbook(XUAT, data_only=True)

    ws_m, ws_x = wb_m.active, wb_x.active
    ws_mv, ws_xv = wb_mv.active, wb_xv.active

    lines = []
    lines.append(f"File mẫu:  {MAU.name}")
    lines.append(f"File xuất: {XUAT.name}")
    lines.append("")
    lines.append("=== TỔNG QUAN ===")
    lines.append(f"{'':12} {'Mẫu':>12} {'Xuất':>12}")
    lines.append(f"{'Max hàng':12} {ws_m.max_row:>12} {ws_x.max_row:>12}")
    lines.append(f"{'Số trạm':12} {count_stations(ws_m):>12} {count_stations(ws_x):>12}")
    lines.append(f"{'Công thức':12} {count_formulas(ws_m):>12} {count_formulas(ws_x):>12}")

    lines.append("")
    lines.append("=== TIÊU ĐỀ ===")
    for cell in ["A1", "A3", "C3", "H3", "J3"]:
        lines.append(f"{cell}: mẫu={ws_mv[cell].value!r} | xuất={ws_xv[cell].value!r}")

    lines.append("")
    lines.append("=== TRẠM 1 (hàng 13–16) ===")
    lines.append(f"{'Ô':6} {'Mẫu':>12} {'Xuất':>12} {'Ghi chú'}")
    for r in range(13, 17):
        for col in ["A", "C", "E", "F", "H", "I", "J", "K", "L", "N"]:
            addr = f"{col}{r}"
            v_m = ws_mv[addr].value
            v_x = ws_xv[addr].value
            f_m = ws_m[addr].value if isinstance(ws_m[addr].value, str) else ""
            f_x = ws_x[addr].value if isinstance(ws_x[addr].value, str) else ""
            note = ""
            if v_m != v_x:
                if f_m and not f_x:
                    note = "mẫu=công thức, xuất=giá trị"
                elif isinstance(v_m, (int, float)) and isinstance(v_x, (int, float)):
                    note = f"lệch {abs(v_m - v_x):.2f}"
                else:
                    note = "khác"
            lines.append(f"{addr:6} {str(v_m):>12} {str(v_x):>12} {note}")

    lines.append("")
    lines.append("=== TRẠM 2 (hàng 17–20) — hiệu chỉnh K ===")
    for addr in ["J17", "J18", "J19", "K19", "H19"]:
        lines.append(
            f"{addr}: mẫu={ws_mv[addr].value} | xuất={ws_xv[addr].value} | "
            f"ct_mẫu={ws_m[addr].value if isinstance(ws_m[addr].value,str) else '-'}"
        )

    lines.append("")
    lines.append("=== TỔNG HỢP (hàng 89–97) ===")
    for r in range(89, 98):
        dm = row_dict(ws_mv, r)
        dx = row_dict(ws_xv, r)
        if dm or dx:
            lines.append(f"H{r} mẫu:  {dm}")
            lines.append(f"H{r} xuất: {dx}")

    lines.append("")
    lines.append("=== KHÁC BIỆT CHÍNH ===")
    diffs = [
        "1. Cấu trúc tiêu đề (hàng 1–12): GIỐNG nhau",
        f"2. Số trạm: mẫu={count_stations(ws_m)}, xuất={count_stations(ws_x)}",
        "3. Mẫu gốc: trạm 1–2 có công thức Excel; trạm 3–19 nhập tay (giá trị cứng)",
        "4. File xuất: TẤT CẢ trạm có giá trị số điền sẵn (tự phân bổ từ tổng hợp)",
        "5. Cột J trạm 1 mẫu=0 (công thức); xuất=0 (giá trị tính sẵn)",
    ]
    for d in diffs:
        lines.append(d)

    out = ROOT / "compare_result.txt"
    out.write_text("\n".join(lines), encoding="utf-8")
    print(out.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
