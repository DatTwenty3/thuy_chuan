"""Đọc dữ liệu mẫu từ So do.xlsx."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

MAU_PATH = Path(__file__).parent / "So do.xlsx"
K1_DEFAULT = 4575.0
K2_DEFAULT = 4475.0


@dataclass
class MauTram:
    kc_sau: float
    kc_truoc: float
    chenh_h: float
    chenh_cao: float
    n_s: float
    n_t: float
    j_mau_s: float
    j_mau_t: float
    offset_h_s: float
    offset_h_t: float
    c_s_base: float
    e_s_base: float


def _n_from_label(label: str | None, k1: float, k2: float) -> float:
    if label == "K1":
        return k1
    if label == "K2":
        return k2
    return k1


def _read_j(cell_value, n: float, h: float, i: float) -> float:
    if cell_value is not None and cell_value != "":
        return float(cell_value)
    return n + h - i


def doc_mau_tu_excel(path: Path | None = None, k1: float = K1_DEFAULT, k2: float = K2_DEFAULT) -> list[MauTram]:
    path = path or MAU_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result: list[MauTram] = []

    for r in range(13, 200, 4):
        if not (ws.cell(r, 1).value and str(ws.cell(r, 1).value).isdigit()):
            break

        r_s, r_t, r_st = r, r + 1, r + 2

        c_s = float(ws.cell(r_s, 3).value or 0)
        c_t = float(ws.cell(r_t, 3).value or 0)
        e_s = float(ws.cell(r_s, 5).value or 0)
        e_t = float(ws.cell(r_t, 5).value or 0)
        h_s = float(ws.cell(r_s, 8).value or 0)
        h_t = float(ws.cell(r_t, 8).value or 0)
        i_s = float(ws.cell(r_s, 9).value or 0)
        i_t = float(ws.cell(r_t, 9).value or 0)

        n_s = _n_from_label(ws.cell(r_s, 12).value, k1, k2)
        n_t = _n_from_label(ws.cell(r_t, 12).value, k1, k2)

        kc_sau = (c_s - c_t) / 10
        kc_truoc = (e_s - e_t) / 10
        chenh_h = h_s - h_t
        chenh_cao = float(
            ws.cell(r_st, 11).value if ws.cell(r_st, 11).value is not None else chenh_h
        )

        mid_c = (c_s + c_t) / 2
        mid_e = (e_s + e_t) / 2
        offset_h_s = h_s - mid_c
        offset_h_t = h_t - mid_e

        j_mau_s = _read_j(ws.cell(r_s, 10).value, n_s, h_s, i_s)
        j_mau_t = _read_j(ws.cell(r_t, 10).value, n_t, h_t, i_t)

        result.append(
            MauTram(
                kc_sau=kc_sau,
                kc_truoc=kc_truoc,
                chenh_h=chenh_h,
                chenh_cao=chenh_cao,
                n_s=n_s,
                n_t=n_t,
                j_mau_s=j_mau_s,
                j_mau_t=j_mau_t,
                offset_h_s=offset_h_s,
                offset_h_t=offset_h_t,
                c_s_base=c_s,
                e_s_base=e_s,
            )
        )

    return result


def tong_mau(mau: list[MauTram]) -> dict[str, float]:
    return {
        "l_sau": sum(t.kc_sau for t in mau),
        "l_truoc": sum(t.kc_truoc for t in mau),
        "h_tong": sum(t.chenh_h for t in mau),
        "chenh_cao_tong": sum(t.chenh_cao for t in mau),
    }
