"""Kiểm thử excel_builder và phan_bo_tram."""

from io import BytesIO

import openpyxl

from excel_builder import DuAn, TongHop, build_workbook, _station_rows
from phan_bo_tram import (
    GIOI_HAN_J,
    GIOI_HAN_OFFSET_H,
    phan_bo_tram,
    tinh_ket_qua_tram,
    tong_kiem_tra,
)


def _build_default():
    tong = TongHop(1794.9, 1798.1, 537, 638)
    tram_list, mau_sel, chenh_cao_list = phan_bo_tram(
        19, 4575, 4475, tong.l_sau, tong.l_truoc, tong.h_tong, tong.i_tong, 100
    )
    wb = build_workbook(
        DuAn(), 19, 4575, 4475, 100, tong, tram_list, mau_sel, chenh_cao_list
    )
    return wb, tram_list, mau_sel, chenh_cao_list, tong


def test_structure(n: int) -> None:
    tong = TongHop(1794.9, 1798.1, 537, 638)
    tram_list, mau_sel, chenh_cao_list = phan_bo_tram(
        n, 4575, 4475, tong.l_sau, tong.l_truoc, tong.h_tong, tong.i_tong, 100
    )
    wb = build_workbook(
        DuAn(), n, 4575, 4475, 100, tong, tram_list, mau_sel, chenh_cao_list
    )
    ws = wb.active
    r_tong = 13 + 4 * n
    r_f = 18 + 4 * n

    assert ws[f"A{r_tong}"].value == "Tổng"
    assert ws[f"H13"].value is not None
    assert ws[f"J13"].value == 0

    s_rows, _ = _station_rows(n)
    expected_sum = "=SUM(" + ",".join(f"C{r}" for r in s_rows) + ")"
    assert ws[f"C{r_f}"].value == expected_sum
    print(f"n={n}: OK")


def test_khop_mau_tram1_2() -> None:
    wb, tram_list, mau_sel, chenh_cao_list, tong = _build_default()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wsv = openpyxl.load_workbook(buf, data_only=False).active

    kq1 = tinh_ket_qua_tram(tram_list[0], mau_sel[0].n_s, mau_sel[0].n_t, 100, 1, chenh_cao_list[0])
    kq2 = tinh_ket_qua_tram(tram_list[1], mau_sel[1].n_s, mau_sel[1].n_t, 100, 2, chenh_cao_list[1])

    assert kq1.chenh_cao == -11, f"K15 expected -11, got {kq1.chenh_cao}"
    assert kq2.j_s == -2, f"J17 expected -2, got {kq2.j_s}"
    assert kq2.j_t == -1, f"J18 expected -1, got {kq2.j_t}"
    assert kq2.chenh_h == 47, f"H19 expected 47, got {kq2.chenh_h}"
    assert kq2.chenh_cao == 48, f"K19 expected 48, got {kq2.chenh_cao}"

    assert wsv["K15"].value == -11
    assert wsv["J17"].value == -2
    assert wsv["J18"].value == -1
    assert wsv["H19"].value == 47
    assert wsv["K19"].value == 48
    print("tram 1-2 khop mau OK")


def test_phan_bo_khop_tong() -> None:
    _, tram_list, mau_sel, chenh_cao_list, _ = _build_default()
    kt = tong_kiem_tra(tram_list, mau_sel, chenh_cao_list, 100)

    assert abs(kt["l_sau"] - 1794.9) < 0.15
    assert abs(kt["l_truoc"] - 1798.1) < 0.15
    assert abs(kt["h_tong"] - 537) < 1
    print(f"phan bo tong: L_sau={kt['l_sau']}, H={kt['h_tong']} OK")


def test_hang4_cot_e_tat_ca_tram() -> None:
    wb, _, _, _, _ = _build_default()
    ws = wb.active

    assert isinstance(ws["E16"].value, (int, float)), "E16 phai la gia tri so"
    assert ws["E20"].value == "=E16+C20"

    for tram in range(3, 20):
        r_delta = 13 + (tram - 1) * 4 + 3
        prev_r = r_delta - 4
        expected = f"=E{prev_r}+C{r_delta}"
        actual = ws[f"E{r_delta}"].value
        assert actual == expected, f"tram {tram}: E{r_delta} expected {expected}, got {actual}"

    print("hang4 cot E tat ca tram OK")


def test_rang_buoc_hang4() -> None:
    _, tram_list, mau_sel, chenh_cao_list, _ = _build_default()

    for i, doc in enumerate(tram_list, start=1):
        mt = mau_sel[i - 1]
        kq = tinh_ket_qua_tram(doc, mt.n_s, mt.n_t, 100, i, chenh_cao_list[i - 1])
        assert abs(kq.sai_so_h_s) <= GIOI_HAN_OFFSET_H, f"tram {i} sai_h_s={kq.sai_so_h_s}"
        assert abs(kq.sai_so_h_t) <= GIOI_HAN_OFFSET_H, f"tram {i} sai_h_t={kq.sai_so_h_t}"
        assert abs(kq.j_s) <= GIOI_HAN_J, f"tram {i} j_s={kq.j_s}"
        assert abs(kq.j_t) <= GIOI_HAN_J, f"tram {i} j_t={kq.j_t}"

    print("rang buoc hang4 OK")


if __name__ == "__main__":
    test_structure(19)
    test_structure(5)
    test_khop_mau_tram1_2()
    test_hang4_cot_e_tat_ca_tram()
    test_phan_bo_khop_tong()
    test_rang_buoc_hang4()
    print("ALL TESTS PASSED")
